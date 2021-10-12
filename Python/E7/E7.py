from pyhunter import PyHunter
from openpyxl import Workbook
import getpass

'#Darian Michelle Orona Aguilar'
'#Gerson Hiram Alcocer Jimenez'
'#Carlos Adrian Soto Serna'
'#Edwin Soto Arregolitia'


def Busqueda(organizacion):
    resultado = hunter.domain_search(company=organizacion,
                                     limit=1,
                                     emails_type='personal')
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    hoja["A1"] = "Valores"
    hoja["B1"] = "Resultados obtenidos"
    count = 2
    for x, y in datosEncontrados.items():
        if type(y) == list:
            for i in y:
                for a, b in i.items():
                    if type(b) == list:
                        for j in b:
                            for c, d in j.items():
                                hoja.cell(count, 1, c)
                                hoja.cell(count, 2, str(d))
                                count += 1
                    else:
                        hoja.cell(count, 1, a)
                        hoja.cell(count, 2, str(b))
                        count += 1
        else:
            hoja.cell(count, 1, x)
            hoja.cell(count, 2, str(y))
            count += 1
    libro.save("Hunter" + organizacion + ".xlsx")

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados, orga)

